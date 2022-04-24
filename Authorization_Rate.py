import csv
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl import *
from datetime import date
from datetime import *
import dateutil.relativedelta as REL


def main():
    count = 0
    
    wb1 = openpyxl.load_workbook("WP RO Analysis (Adam).xlsx")
    sheet = wb1["WP Data - Raw"]

    #CREATE AND POPULATE DICTIONARIES THAT ARE USED TO FILL THE SHEETS DOWN LOWER
    myDict = {}
    autoShipDict = {}
    consecTracker = {}
    #get data from worldpay file and put relevant data into a dictionary
    #relevant data by order = order number, payment type, final event type, number of failed attempts, final date/time, name, country, Auto-Ship ID, consecutive recurring number
    for row in sheet.iter_rows(min_row=2, max_col=18):
        orderNum = int(row[5].value)
        paymentType = row[3].value
        eventType = row[4].value
        if eventType == "REFUSED":
            failedAttempts = 1
        else:
            failedAttempts = 0
        date_time = row[7].value
        name = row[16].value
        country = row[15].value
        autoShip = str(row[6].value)
        consecNum = 0
        Merchant = row[0].value

        if autoShip != "#N/A" and autoShip != "0":
            if autoShip not in consecTracker:
                consecTracker.update({autoShip:[orderNum]})

            elif autoShip in consecTracker:
                orders2 = consecTracker.get(autoShip)
                exist_count = orders2.count(orderNum)
                if exist_count == 0:
                    orders2.append(orderNum)
                    consecTracker.update({autoShip: orders2})


        if orderNum not in myDict:
            if eventType == "REFUSED" or eventType == "AUTHORISED":
                if autoShip in consecTracker:
                    consecNum = len(consecTracker.get(autoShip))
                myDict.update({orderNum:[paymentType, eventType, failedAttempts, date_time, name, country, autoShip, consecNum, Merchant]})

        elif orderNum in myDict:
            values = myDict.get(orderNum)
            if eventType == "AUTHORISED" and values[1] == "REFUSED":
                values[1] = eventType
                values[3] = date_time
                myDict.update({orderNum: values})
                
            elif eventType == "REFUSED" and values[1] == "REFUSED":
                num = values[2]
                num += 1
                values[2] = num
                values[3] = date_time
                myDict.update({orderNum: values})

            elif eventType == "REFUSED" and values[1] == "AUTHORISED":
                num = values[2]
                num += 1
                values[2] = num
                myDict.update({orderNum: values})


    #ORDERS SHEET TAB
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    ws.cell(row = 1, column = 1, value = "Order Number")
    ws.cell(row = 1, column = 2, value = "Event Type")
    ws.cell(row = 1, column = 3, value = "Num of Failed Attempts")
    ws.cell(row = 1, column = 4, value = "Date-Time")
    ws.cell(row = 1, column = 5, value = "Merchant")
    ws.cell(row = 1, column = 6, value = "Auto-Ship ID")
    ws.cell(row = 1, column = 7, value = "Consecutive Num")
    ws.cell(row = 1, column = 8, value = "Name")
    ws.cell(row = 1, column = 9, value = "Country")
    ws.cell(row = 1, column = 10, value = "payment Type")

    rowIndex = 2
    nameDict = {}
    totals = [0, 0, 0]
    totalsR1 = [0, 0, 0]
    totalsR2 = [0, 0, 0]
    for key in myDict:
        values = myDict.get(key)
        orderNum = key
        paymentType = values[0]
        eventType = values[1]
        failedAttempts = values[2]
        date_time = values[3]
        name = values[4]
        country = values[5]
        autoShip = values[6]
        consecNum = values[7]
        merchant = values[8]
        
        ws.cell(row = rowIndex, column = 1, value = orderNum)
        ws.cell(row = rowIndex, column = 2, value = eventType)
        ws.cell(row = rowIndex, column = 3, value = failedAttempts)
        ws.cell(row = rowIndex, column = 4, value = date_time)
        ws.cell(row = rowIndex, column = 5, value = merchant)
        ws.cell(row = rowIndex, column = 6, value = autoShip)
        ws.cell(row = rowIndex, column = 7, value = consecNum)   
        ws.cell(row = rowIndex, column = 8, value = name)
        ws.cell(row = rowIndex, column = 9, value = country)
        ws.cell(row = rowIndex, column = 10, value = paymentType)
        
        rowIndex += 1

        #get info ready for 'AutoShip' tab
        if autoShip != "#N/A" and autoShip != "0":
            if autoShip not in autoShipDict:
                if eventType == "REFUSED":
                    autoShipDict.update({autoShip:[1, merchant, name, orderNum, eventType]})
                if eventType == "AUTHORISED":
                    autoShipDict.update({autoShip:[0, merchant, name, orderNum, eventType]})

            elif autoShip in autoShipDict:
                orders2 = autoShipDict.get(autoShip)
                if merchant != orders2[1]:
                    orders2[1] = merchant
                exist_count = orders2.count(orderNum)
                if exist_count == 0:
                    orders2.append(orderNum)
                    orders2.append(eventType)
                    autoShipDict.update({autoShip: orders2})
                    

        #get info ready for 'Totals' Tab
        if eventType == "REFUSED":
            failed = totals[0]
            total = totals[1]
            failed += 1
            total += 1
            percent = str(int((1-(failed/total))*100))+"%"
            totals[0] = failed
            totals[1] = total
            totals[2] = percent
            
        elif eventType == "AUTHORISED":
            failed = totals[0]
            total = totals[1]
            total += 1
            percent = str(int((1-(failed/total))*100))+"%"
            totals[0] = failed
            totals[1] = total
            totals[2] = percent




    #AUTOSHIP SHEET TAB
    wb.create_sheet(index=1, title='AutoShip')
    ws2 = wb['AutoShip']
    ws2.cell(row = 1, column = 1, value = "AutoShip ID")
    ws2.cell(row = 1, column = 2, value = "Did First Fail?")
    ws2.cell(row = 1, column = 3, value = "Number of Orders (Total)")
    ws2.cell(row = 1, column = 4, value = "Number of Failed Orders")
    ws2.cell(row = 1, column = 5, value = "Success Rate %")
    ws2.cell(row = 1, column = 6, value = "Name")
    ws2.cell(row = 1, column = 7, value = "Merchant")

    rowIndex = 2
    for key in autoShipDict:
        values = autoShipDict.get(key)
        autoID = key
        firstFailedNum = values[0]
        if firstFailedNum == 1:
            firstFailed = "Yes"
        if firstFailedNum == 0:
            firstFailed = "No"  
        merchant = values[1]
        name = values[2]
        numOfOrders = (len(values)-3)/2
        numOfFailed = values.count("REFUSED")
        successRate = int((1-(numOfFailed/numOfOrders))*100)
        
        ws2.cell(row = rowIndex, column = 1, value = int(autoID))
        ws2.cell(row = rowIndex, column = 2, value = firstFailed)
        ws2.cell(row = rowIndex, column = 3, value = numOfOrders)
        ws2.cell(row = rowIndex, column = 4, value = numOfFailed)
        ws2.cell(row = rowIndex, column = 5, value = successRate)
        ws2.cell(row = rowIndex, column = 6, value = name)
        ws2.cell(row = rowIndex, column = 7, value = merchant)

        rowIndex += 1



    
    '''
    #TOTALS SHEET TAB
    wb.create_sheet(index=2, title='Totals')
    ws3 = wb['Totals']
    ws3.cell(row = 1, column = 1, value = "Num of Failed Orders")
    ws3.cell(row = 1, column = 2, value = "Num of Total Orders")
    ws3.cell(row = 1, column = 3, value = "Success rate %")
    colIndex = 1
    for item in totals:
        ws3.cell(row = 2, column = colIndex, value = item)
        colIndex += 1
        
    ws3.cell(row = 6, column = 1, value = "Merchant")
    ws3.cell(row = 6, column = 2, value = "% Failed first, passed a future one")
    ws3.cell(row = 6, column = 3, value = "% Passed first, failed a future one")
    ws3.cell(row = 6, column = 4, value = "% With all Fails (0% success)")
    ws3.cell(row = 6, column = 5, value = "% With no Fails (100% success)")
    colIndex = 1
    merchantDict = {"RAININTBVEUR":["BV EUR", 0, 0, 0], "RAININTBVEURREC":["BV EUR REC", 0, 0, 0],
                    "RAININTUSD":["USD", 0, 0, 0], "RAININTUSDREC":["USD REC", 0, 0, 0],
                    "RAININTUSHKD":["HKD", 0, 0, 0], "RAININTUSHKDREC":["HKD REC", 0, 0, 0],
                    "RAININTUSMXN":["MXN", 0, 0, 0], "RAININTUSMXNREC":["MXN REC", 0, 0, 0],
                    "RAININTUSSGD":["SGD", 0, 0, 0], "RAININTUSSGDREC":["SGD REC", 0, 0, 0]}
    numOfAutoShipOrders = len(autoShipDict)
    for key in autoShipDict:
        #get info from AutShip Dictionary
        values = autoShipDict.get(key)
        didFirstFail = values[0]
        merchant = values[1]
        numOfOrders = (len(values)-3)/2
        numOfFailed = values.count("REFUSED")
        percent
        

        #update info in merchant Dictionary
        merchInfo = merchantDict.get(merchant)
        numOfFirstFails = merchInfo[1]
        numOfFailedOrders = merchInfo[2]
        numOfTotalOrders = merchInfo[3]
        
        numOfFirstFails = numOfFirstFails + didFirstFail
        numOfFailedOrders = numOfFailedOrders + numOfFailed
        numOfTotalOrders = numOfTotalOrders + numOfOrders

        merchInfo[1] = numOfFirstFails
        merchInfo[2] = numOfFailedOrders
        merchInfo[3] = numOfTotalOrders
        merchantDict.update({merchant: merchInfo})
    '''
            

    wb.save("NEW WP OR Analysis Report"+ ".xlsx")
    print("Done")

main()
